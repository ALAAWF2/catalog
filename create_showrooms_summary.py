import pandas as pd
import json
import os
import re

# Profiles retrieved from Supabase
profiles_json = [
    {"email":"mujanpark@orangebedbath.com","warehouse":"Warehouse","mall":"43-Mujan Park","role":"user"},
    {"email":"riyadh.gallery@orangebedbath.com","warehouse":"warehouse riyadh","mall":"45- Riyadh Gallery Mall","role":"user"},
    {"email":"dhahran.mall@orangebedbath.com","warehouse":"warehouse riyadh","mall":"27-Dhahran Mall khobar","role":"user"},
    {"email":"azizmall2@orangebedbath.com","warehouse":"Warehouse","mall":"56- Aziz Mall 2","role":"user"},
    {"email":"makkah.mall@orangebedbath.com","warehouse":"Warehouse","mall":"08-Makkah Mall","role":"user"},
    {"email":"kamal.taif@orangebedbath.com","warehouse":"Warehouse","mall":"14-Al Kamal Mall","role":"user"},
    {"email":"nakheel.plaza@orangebedbath.com","warehouse":"warehouse riyadh","mall":"47-Al-Nakheel Plaza","role":"user"},
    {"email":"arab.jed@orangebedbath.com","warehouse":"Warehouse","mall":"07-Arab Mall","role":"user"},
    {"email":"dana.yanbu@orangebedbath.com","warehouse":"Warehouse","mall":"24-Yanbu Dana Mall","role":"user"},
    {"email":"souq7@orangebedbath.com","warehouse":"Warehouse","mall":"57-Sauq7","role":"user"},
    {"email":"basateenmall@orangebedbath.com","warehouse":"Warehouse","mall":"53-Al Basateen Mall","role":"user"},
    {"email":"alia.madina@orangebedbath.com","warehouse":"Warehouse","mall":"23-Alia Mall Madinah","role":"user"},
    {"email":"yasmin.jed@orangebedbath.com","warehouse":"Warehouse","mall":"13-Al-Yasmin Mall","role":"user"},
    {"email":"tasan.makkah@orangebedbath.com","warehouse":"Warehouse","mall":"20-Sitten Street Makkah","role":"user"},
    {"email":"albahamall@orangebedbath.com","warehouse":"Warehouse","mall":"52-Al_Baha Mall","role":"user"},
    {"email":"othaim.hail@orangebedbath.com","warehouse":"warehouse riyadh","mall":"19-Hail Othaim Mall","role":"user"},
    {"email":"khaleej.riyadh@orangebedbath.com","warehouse":"warehouse riyadh","mall":"46-Khaleej Mall Riyadh","role":"user"},
    {"email":"othaim.ehsa@orangebedbath.com","warehouse":"warehouse riyadh","mall":"16-Ehsa Othaim Mall","role":"user"},
    {"email":"tala.riyadh@orangebedbath.com","warehouse":"warehouse riyadh","mall":"30-Tala Mall Riyadh","role":"user"},
    {"email":"salam.jed@orangebedbath.com","warehouse":"Warehouse","mall":"09-Al-Salam Mall","role":"user"},
    {"email":"othaim.riyad@orangebedbath.com","warehouse":"warehouse riyadh","mall":"15-Riyadh Othaim Mall","role":"user"},
    {"email":"nakheel.riyadh@orangebedbath.com","warehouse":"warehouse riyadh","mall":"29-Al Nakheel Mall Riyadh","role":"user"},
    {"email":"jouri.taif@orangebedbath.com","warehouse":"Warehouse","mall":"11-Jouri Mall","role":"user"},
    {"email":"alrabie@orangebedbath.com","warehouse":"warehouse riyadh","mall":"1115-Alrabie Mall","role":"user"},
    {"email":"alrashid.abha@orangebedbath.com","warehouse":"Warehouse","mall":"21-Abha Al_Rashid Mall New","role":"user"},
    {"email":"nakheel.dammam@orangebedbath.com","warehouse":"warehouse riyadh","mall":"28-Al Nakheel Mall Dammam","role":"user"},
    {"email":"meem.plaza@orangebedbath.com","warehouse":"warehouse riyadh","mall":"50-Meem Plaza Riyadh","role":"user"},
    {"email":"riyadh.park@orangebedbath.com","warehouse":"warehouse riyadh","mall":"38-Al_Riyadh Park","role":"user"},
    {"email":"malgha@orangebedbath.com","warehouse":"warehouse riyadh","mall":"1114-Malgha Mall","role":"user"},
    {"email":"thevillage@orangebedbath.com","warehouse":"Warehouse","mall":"54-THE VILLAGE","role":"user"},
    {"email":"salam.riyadh@orangebedbath.com","warehouse":"warehouse riyadh","mall":"39-Salam MAll Riyadh","role":"user"},
    {"email":"dareen.dammam@orangebedbath.com","warehouse":"warehouse riyadh","mall":"42-Dareen Mall Dammam","role":"user"},
    {"email":"andalos.jed@orangebedbath.com","warehouse":"Warehouse","mall":"04-Andalos Mall","role":"user"},
    {"email":"othaim.arar@orangebedbath.com","warehouse":"Warehouse","mall":"17-Arar Othaim Mall","role":"user"},
    {"email":"haifa.jed@orangebedbath.com","warehouse":"Warehouse","mall":"05-Haifa Mall","role":"user"},
    {"email":"alnoor.madinah@orangebedbath.com","warehouse":"Warehouse","mall":"26-Al-Noor Mall Madinah","role":"user"},
    {"email":"khamis.avenue@orangebedbath.com","warehouse":"Warehouse","mall":"41-Khamis Avenue","role":"user"},
    {"email":"parkavenue.riyadh@orangebedbath.com","warehouse":"warehouse riyadh","mall":"51-Park Avenue Riyadh","role":"user"},
    {"email":"jubail.mall@orangebedbath.com","warehouse":"warehouse riyadh","mall":"36-Al jubail Mall","role":"user"},
    {"email":"atyaf.riyad@orangebedbath.com","warehouse":"warehouse riyadh","mall":"32-Atyaf Mall Riyadh","role":"user"},
    {"email":"tabuk.park@orangebedbath.com","warehouse":"Warehouse","mall":"22-Tabuk Park","role":"user"},
    {"email":"hayat.riyadh@orangebedbath.com","warehouse":"warehouse riyadh","mall":"40-Hayat Mall Riyadh","role":"user"},
    {"email":"hamra.riyadh@orangebedbath.com","warehouse":"warehouse riyadh","mall":"12-Al_Hamra Mall","role":"user"},
    {"email":"rabwa.riyad@orangebedbath.com","warehouse":"warehouse riyadh","mall":"25-Rabwa Othaim Mall","role":"user"},
    {"email":"aljouf.center@orangebedbath.com","warehouse":"Warehouse","mall":"44-Al-Jouf Center","role":"user"},
    {"email":"lavanda.park@orangebedbath.com","warehouse":"Warehouse","mall":"1906-LAVANDA PARK","role":"user"},
    {"email":"redsea.jed@orangebedbath.com","warehouse":"Warehouse","mall":"06-Red Sea Mall","role":"user"},
    {"email":"jabalomar@orangebedbath.com","warehouse":"Warehouse","mall":"55- Jabl Omar","role":"user"},
    {"email":"alahsamall@orangebedbath.com","warehouse":"warehouse riyadh","mall":"49-AlAhsa Mall","role":"user"},
    {"email":"khayyat.jed@orangebedbath.com","warehouse":"Warehouse","mall":"18-Al_Khayyat Center","role":"user"},
    {"email":"jeddah.park@orangebedbath.com","warehouse":"Warehouse","mall":"48 - Jeddah Park","role":"user"}
]

# Manager mappings
managers = {
    "obieda.sebaee@orangebedbath.com": {
        "name": "عبيدة السباعي",
        "outlets": ["04-Andalos Mall", "09-Al-Salam Mall", "18-Al_Khayyat Center", "53-Al Basateen Mall", "57-Sauq7"]
    },
    "mehyar.s@orangebedbath.com": {
        "name": "مهيار",
        "outlets": ["05-Haifa Mall", "06-Red Sea Mall", "07-Arab Mall", "13-Al-Yasmin Mall", "48 - Jeddah Park", "54-THE VILLAGE", "56- Aziz Mall 2"]
    },
    "radwan@orangebedbath.com": {
        "name": "رضوان عطيوي",
        "outlets": ["08-Makkah Mall", "11-Jouri Mall", "14-Al Kamal Mall", "20-Sitten Street Makkah", "55- Jabl Omar"]
    },
    "m.kello@orangebedbath.com": {
        "name": "محمد كلو",
        "outlets": ["1114-Malgha Mall", "1115-Alrabie Mall", "12-Al_Hamra Mall", "29-Al Nakheel Mall Riyadh", "32-Atyaf Mall Riyadh", "38-Al_Riyadh Park", "45- Riyadh Gallery Mall", "46-Khaleej Mall Riyadh", "51-Park Avenue Riyadh"]
    },
    "abd.serdah@orangebedbath.com": {
        "name": "عبدالله السرداح",
        "outlets": ["15-Riyadh Othaim Mall", "19-Hail Othaim Mall", "25-Rabwa Othaim Mall", "30-Tala Mall Riyadh", "39-Salam MAll Riyadh", "40-Hayat Mall Riyadh", "47-Al-Nakheel Plaza", "50-Meem Plaza Riyadh"]
    },
    "jihad@orangebedbath.com": {
        "name": "جهاد ايوبي",
        "outlets": ["16-Ehsa Othaim Mall", "27-Dhahran Mall khobar", "28-Al Nakheel Mall Dammam", "36-Al jubail Mall", "42-Dareen Mall Dammam", "49-AlAhsa Mall"]
    },
    "bakr@orangebedbath.com": {
        "name": "بكر",
        "outlets": ["17-Arar Othaim Mall", "22-Tabuk Park", "23-Alia Mall Madinah", "24-Yanbu Dana Mall", "26-Al-Noor Mall Madinah", "44-Al-Jouf Center"]
    },
    "amani.a@orangebedbath.com": {
        "name": "اماني عسيري",
        "outlets": ["1906-LAVANDA PARK", "21-Abha Al_Rashid Mall New", "41-Khamis Avenue", "43-Mujan Park", "52-Al_Baha Mall"]
    }
}

# Helper to extract code from showroom name (e.g. "04-Andalos Mall" -> "04", "48 - Jeddah Park" -> "48")
def extract_showroom_code(name):
    m = re.match(r'^(\d+)\s*-\s*', name)
    if m:
        return m.group(1)
    # also try just digits at the start
    m = re.match(r'^(\d+)', name)
    if m:
        return m.group(1)
    return None

# Load the master showrooms from Table5
if not os.path.exists("mapping.xlsx"):
    print("Error: mapping.xlsx not found.")
    exit(1)

df_table5 = pd.read_excel("mapping.xlsx", sheet_name="Table5")

# We want to map: store number, outlet name, showroom email, warehouse code, warehouse arabic, manager name, manager email
rows = []
for idx, row in df_table5.iterrows():
    outlet_name = str(row.get("outlet")).strip()
    store_num = str(row.get("store number")).strip()
    
    if outlet_name == 'nan' or not outlet_name:
        continue
        
    # Extract code for robust prefix-based matching
    outlet_code = extract_showroom_code(outlet_name)
    
    # Find warehouse and user email from user profiles
    warehouse_code = ""
    showroom_email = ""
    
    # Try match by code or exact name in profiles
    matched_profile = None
    for p in profiles_json:
        p_mall = p["mall"].strip()
        p_code = extract_showroom_code(p_mall)
        if (p_code and outlet_code and p_code == outlet_code) or (p_mall.lower() == outlet_name.lower()):
            matched_profile = p
            break
            
    if matched_profile:
        warehouse_code = matched_profile["warehouse"]
        showroom_email = matched_profile["email"]
    else:
        # Fallback based on name for warehouse
        name_lower = outlet_name.lower()
        if any(x in name_lower for x in ["riyadh", "الرياض", "dammam", "الدمام", "khobar", "الخبر", "hail", "حائل", "ehsa", "الأحساء", "ahsa", "jubail", "الجبيل", "qassim", "القصيم", "buraidah", "بريدة", "nakheel", "عنيزة"]):
            warehouse_code = "warehouse riyadh"
        else:
            warehouse_code = "Warehouse"
            
    # Warehouse Arabic translation
    warehouse_ar = "مستودع جدة (الرئيسي)"
    if warehouse_code == "warehouse riyadh":
        warehouse_ar = "مستودع الرياض (الوسطى)"
        
    # Find manager using code or exact name matching
    mgr_name = ""
    mgr_email = ""
    for email, data in managers.items():
        matched_mgr = False
        for out in data["outlets"]:
            out_code = extract_showroom_code(out)
            if (out_code and outlet_code and out_code == outlet_code) or (out.lower() == outlet_name.lower()):
                matched_mgr = True
                break
        if matched_mgr:
            mgr_name = data["name"]
            mgr_email = email
            break
            
    # Remove duplicates from virtual system names
    if outlet_name.lower() in ["platforms", "00-gifts"] or "warehouse" in outlet_name.lower():
        mgr_name = "مستودع رئيسي / قناة نظام"
        
    rows.append({
        "كود المعرض (Store Number)": store_num,
        "اسم المعرض": outlet_name,
        "إيميل المعرض": showroom_email if showroom_email else (f"{outlet_name.lower().replace(' ', '').replace('-', '')}@orangebedbath.com" if matched_profile else "System/Virtual"),
        "كود المستودع المربوط (D365)": warehouse_code,
        "المستودع الذي يرى بضاعته": warehouse_ar,
        "المدير الإقليمي المسؤول": mgr_name if mgr_name else "غير محدد",
        "إيميل المدير": mgr_email if mgr_email else "-"
    })

df_result = pd.DataFrame(rows)

# Drop any exact duplicates
df_result = df_result.drop_duplicates(subset=["كود المعرض (Store Number)", "اسم المعرض"])

# Sort by code
df_result = df_result.sort_values(by="كود المعرض (Store Number)")

# Export to Excel with beautiful styling
output_file = "showrooms_summary.xlsx"
writer = pd.ExcelWriter(output_file, engine='openpyxl')
df_result.to_excel(writer, sheet_name="المعارض والمستودعات والمدراء", index=False)

# Styling with openpyxl
workbook = writer.book
worksheet = writer.sheets["المعارض والمستودعات والمدراء"]

# Styling properties
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

# Apply header styles
for col_num in range(1, len(df_result.columns) + 1):
    cell = worksheet.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# Enable grid lines
worksheet.views.sheetView[0].showGridLines = True

# Style data cells
for row_num in range(2, len(df_result) + 2):
    for col_num in range(1, len(df_result.columns) + 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.border = thin_border
        
        # Alignment & font
        if col_num in [1, 4]: # store number, warehouse code
            cell.alignment = center_align
        elif col_num in [3, 7]: # emails
            cell.alignment = left_align
        else: # arabic texts
            cell.alignment = right_align
            
        cell.font = Font(name="Calibri", size=11)
        
        # Alternate row fill
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")

# Auto-adjust column widths
for col in worksheet.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = col[0].column_letter
    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

writer.close()
print(f"Excel file created successfully: {output_file}")
